import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# Step 1: Read the CSV data
data = pd.read_csv('train_features_scaled.csv')

# Step 2: Round numerical data for simplification (e.g., to 2 decimal places)
rounded_data = data.round(2)

# Step 3: Select a subset of columns for better visualization
columns_to_display = [
    'file_name', 'label',
    'mfcc_mean_1', 'mfcc_std_1',
    'chroma_mean_1', 'chroma_std_1',
    'spec_contrast_mean_1', 'spec_contrast_std_1',
    'zcr_mean', 'zcr_std'
]

# Create a new DataFrame with selected columns
display_data = rounded_data[columns_to_display]

# Step 4: Define colors for columns (aRGB format without '#')
column_colors = [
    'FFFFCCCC',  # Light Red
    'FFCCFFCC',  # Light Green
    'FFCCCCFF',  # Light Blue
    'FFFFFFCC',  # Light Yellow
    'FFCCFFFF',  # Light Cyan
    'FFFFCCFF'   # Light Magenta
]  # Add more colors if more columns are present

# Create a new Excel workbook and select the active worksheet
wb = Workbook()
ws = wb.active
ws.title = "Colorful Table"

# Define header style
header_fill = PatternFill(start_color='40466E', end_color='40466E', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)
header_alignment = Alignment(horizontal='center', vertical='center')

# Write the header and apply column-based colors
for col_num, column_title in enumerate(display_data.columns, 1):
    cell = ws.cell(row=1, column=col_num, value=column_title)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

    # Apply the color to the entire column
    column_color = column_colors[(col_num - 1) % len(column_colors)]
    for row_num in range(2, len(display_data) + 2):  # Data starts at row 2
        data_cell = ws.cell(row=row_num, column=col_num, value=display_data.iloc[row_num - 2, col_num - 1])
        data_cell.fill = PatternFill(start_color=column_color, end_color=column_color, fill_type='solid')
        data_cell.alignment = Alignment(horizontal='center', vertical='center')

# Adjust column widths for better readability
for column_cells in ws.columns:
    max_length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None)
    adjusted_width = (max_length + 2)
    column_letter = column_cells[0].column_letter
    ws.column_dimensions[column_letter].width = adjusted_width

# Step 5: Save the workbook to an Excel file
output_file = 'colorful_table_by_columns.xlsx'
wb.save(output_file)

print(f"Colorful Excel table with column-based colors has been generated and saved as '{output_file}'.")
